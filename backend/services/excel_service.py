"""
services/excel_service.py

Service untuk generate laporan Excel (.xlsx) dari data transaksi.
File Excel yang dihasilkan bisa dikirim via WhatsApp sebagai attachment.
"""

import os
import tempfile
from datetime import datetime
from typing import List

from loguru import logger  # type: ignore
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

from models.transaction import Transaction
from utils.number_formatter import format_rupiah


class ExcelService:
    """
    Service untuk membuat file Excel laporan keuangan.
    
    Output: file .xlsx yang siap dikirim sebagai dokumen WhatsApp.
    """

    # Styling constants
    _HEADER_FILL = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")  # Purple
    _HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    _INCOME_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # Light green
    _EXPENSE_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Light red
    _TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="6B21A8")
    _SUBTITLE_FONT = Font(name="Calibri", size=11, color="64748B")
    _SUMMARY_FONT = Font(name="Calibri", bold=True, size=11)
    _INCOME_FONT = Font(name="Calibri", bold=True, color="16A34A")
    _EXPENSE_FONT = Font(name="Calibri", bold=True, color="DC2626")
    _THIN_BORDER = Border(
        bottom=Side(style="thin", color="E2E8F0"),
    )
    _CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    _RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
    _WRAP_ALIGN = Alignment(wrap_text=True, vertical="center")

    async def generate_daily_report(
        self,
        transactions: List[Transaction],
        phone_number: str,
        report_date: datetime,
        total_income: float,
        total_expense: float,
        net_profit: float,
    ) -> str:
        """
        Generate file Excel untuk laporan harian.
        
        Returns:
            Path ke file .xlsx yang sudah dibuat.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Harian"
        
        # Set column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 12

        row = 1

        # Title Section
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        title_cell = ws.cell(row=row, column=1, value="📊 Laporan Keuangan Harian")
        title_cell.font = self._TITLE_FONT
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        subtitle = ws.cell(row=row, column=1, value=f"Tanggal: {report_date.strftime('%d %B %Y')}  |  No. HP: {phone_number}")
        subtitle.font = self._SUBTITLE_FONT
        row += 2

        # Summary Section
        summary_data = [
            ("💰 Total Pemasukan:", format_rupiah(total_income), self._INCOME_FONT),
            ("💸 Total Pengeluaran:", format_rupiah(total_expense), self._EXPENSE_FONT),
            ("📈 Laba Bersih:", format_rupiah(net_profit), self._SUMMARY_FONT),
            ("📝 Jumlah Transaksi:", str(len(transactions)), self._SUMMARY_FONT),
        ]
        for label, value, font in summary_data:
            ws.cell(row=row, column=1, value=label).font = self._SUMMARY_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            val_cell = ws.cell(row=row, column=4, value=value)
            val_cell.font = font
            row += 1

        row += 1  # Spacer

        # Table Headers
        headers = ["No", "Deskripsi", "Tipe", "Jumlah (Rp)", "Waktu"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self._HEADER_FONT
            cell.fill = self._HEADER_FILL
            cell.alignment = self._CENTER_ALIGN
        row += 1

        # Table Data
        for idx, tx in enumerate(transactions, 1):
            is_income = tx.transaction_type == "income"
            fill = self._INCOME_FILL if is_income else self._EXPENSE_FILL
            
            no_cell = ws.cell(row=row, column=1, value=idx)
            no_cell.alignment = self._CENTER_ALIGN
            no_cell.border = self._THIN_BORDER

            desc_cell = ws.cell(row=row, column=2, value=tx.description)
            desc_cell.alignment = self._WRAP_ALIGN
            desc_cell.border = self._THIN_BORDER

            type_cell = ws.cell(row=row, column=3, value="Pemasukan" if is_income else "Pengeluaran")
            type_cell.fill = fill
            type_cell.alignment = self._CENTER_ALIGN
            type_cell.border = self._THIN_BORDER

            amount_str = f"+{format_rupiah(tx.amount)}" if is_income else f"-{format_rupiah(tx.amount)}"
            amount_cell = ws.cell(row=row, column=4, value=amount_str)
            amount_cell.font = self._INCOME_FONT if is_income else self._EXPENSE_FONT
            amount_cell.alignment = self._RIGHT_ALIGN
            amount_cell.border = self._THIN_BORDER

            time_cell = ws.cell(row=row, column=5, value=tx.transaction_date.strftime("%H:%M"))
            time_cell.alignment = self._CENTER_ALIGN
            time_cell.border = self._THIN_BORDER

            row += 1

        # Footer
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        footer = ws.cell(row=row, column=1, value="Dibuat otomatis oleh FinanceBot UMKM — Chatet")
        footer.font = Font(name="Calibri", italic=True, color="94A3B8", size=9)

        # Save to temp file
        date_str = report_date.strftime("%Y%m%d")
        filename = f"Laporan_Harian_{date_str}.xlsx"
        filepath = os.path.join(tempfile.gettempdir(), filename)
        wb.save(filepath)
        logger.info(f"📄 Excel daily report saved: {filepath}")
        return filepath

    async def generate_monthly_report(
        self,
        transactions: List[Transaction],
        phone_number: str,
        year: int,
        month: int,
        month_name: str,
        total_income: float,
        total_expense: float,
        total_debt: float,
        net_profit: float,
    ) -> str:
        """
        Generate file Excel untuk laporan bulanan.
        
        Returns:
            Path ke file .xlsx yang sudah dibuat.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Bulanan"

        # Set column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 14

        row = 1

        # Title Section
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        title_cell = ws.cell(row=row, column=1, value="📊 Laporan Keuangan Bulanan")
        title_cell.font = self._TITLE_FONT
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        subtitle = ws.cell(row=row, column=1, value=f"Periode: {month_name} {year}  |  No. HP: {phone_number}")
        subtitle.font = self._SUBTITLE_FONT
        row += 2

        # Summary Section
        summary_data = [
            ("💰 Total Pemasukan:", format_rupiah(total_income), self._INCOME_FONT),
            ("💸 Total Pengeluaran:", format_rupiah(total_expense), self._EXPENSE_FONT),
            ("💳 Total Hutang:", format_rupiah(total_debt), self._SUMMARY_FONT),
            ("📈 Laba Bersih:", format_rupiah(net_profit), self._SUMMARY_FONT),
            ("📝 Jumlah Transaksi:", str(len(transactions)), self._SUMMARY_FONT),
        ]
        for label, value, font in summary_data:
            ws.cell(row=row, column=1, value=label).font = self._SUMMARY_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            val_cell = ws.cell(row=row, column=4, value=value)
            val_cell.font = font
            row += 1

        row += 1  # Spacer

        # Table Headers
        headers = ["No", "Deskripsi", "Tipe", "Jumlah (Rp)", "Tanggal"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self._HEADER_FONT
            cell.fill = self._HEADER_FILL
            cell.alignment = self._CENTER_ALIGN
        row += 1

        # Table Data
        for idx, tx in enumerate(transactions, 1):
            is_income = tx.transaction_type == "income"
            fill = self._INCOME_FILL if is_income else self._EXPENSE_FILL

            no_cell = ws.cell(row=row, column=1, value=idx)
            no_cell.alignment = self._CENTER_ALIGN
            no_cell.border = self._THIN_BORDER

            desc_cell = ws.cell(row=row, column=2, value=tx.description)
            desc_cell.alignment = self._WRAP_ALIGN
            desc_cell.border = self._THIN_BORDER

            type_cell = ws.cell(row=row, column=3, value="Pemasukan" if is_income else "Pengeluaran")
            type_cell.fill = fill
            type_cell.alignment = self._CENTER_ALIGN
            type_cell.border = self._THIN_BORDER

            amount_str = f"+{format_rupiah(tx.amount)}" if is_income else f"-{format_rupiah(tx.amount)}"
            amount_cell = ws.cell(row=row, column=4, value=amount_str)
            amount_cell.font = self._INCOME_FONT if is_income else self._EXPENSE_FONT
            amount_cell.alignment = self._RIGHT_ALIGN
            amount_cell.border = self._THIN_BORDER

            date_cell = ws.cell(row=row, column=5, value=tx.transaction_date.strftime("%d/%m/%Y"))
            date_cell.alignment = self._CENTER_ALIGN
            date_cell.border = self._THIN_BORDER

            row += 1

        # Footer
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        footer = ws.cell(row=row, column=1, value="Dibuat otomatis oleh FinanceBot UMKM — Chatet")
        footer.font = Font(name="Calibri", italic=True, color="94A3B8", size=9)

        # Save to temp file
        filename = f"Laporan_Bulanan_{month_name}_{year}.xlsx"
        filepath = os.path.join(tempfile.gettempdir(), filename)
        wb.save(filepath)
        logger.info(f"📄 Excel monthly report saved: {filepath}")
        return filepath
